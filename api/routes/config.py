"""
Config endpoint: presets, tile metadata, decks, option choices.
"""

from __future__ import annotations

import csv
import random
import io
import json
import re
import sys
from pathlib import Path
from typing import Any

from fastapi import APIRouter, File, HTTPException, UploadFile

from api.schemas import (
    CardTemplateCreateRequest,
    CardTemplateEntry,
    CardTemplateUpdateRequest,
    ConfigResponse,
    DeckCreateRequest,
    DeckEntry,
    DeckImportResult,
    DeckUpdateRequest,
    TileMetadataPatchRequest,
    TileStyle,
)
from board_core import (
    BOARD_PRESETS,
    BOARD_SIZE_MAX,
    BOARD_SIZE_MIN,
    GENERATION_MODE_CHOICES,
    TILE_COLORS,
    TILE_METADATA,
    TILE_NAMES,
    TILE_RULES,
    TILE_STYLES,
    TILESET_PRESETS,
    get_tile_metadata,
    set_tile_metadata,
)

router = APIRouter(prefix="/config", tags=["config"])

# In-memory deck store (id -> {name, card_template_ids}). Seeded with defaults or loaded from disk.
_DECKS: dict[str, dict[str, Any]] = {}

# In-memory card template store (id -> {title, body, image_url?, back_text?}).
_CARD_TEMPLATES: dict[str, dict[str, Any]] = {}

_DECKS_INITIALIZED = False


def _get_decks_data_path() -> Path:
    """Writable directory for decks.json. Next to exe when frozen, else project root."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).resolve().parent.parent


def _load_decks_from_disk() -> bool:
    """Load decks and card_templates from decks.json. Replace in-memory store. Return True if file existed and loaded."""
    path = _get_decks_data_path() / "decks.json"
    if not path.exists():
        return False
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        decks = data.get("decks")
        templates = data.get("card_templates")
        if not isinstance(decks, dict):
            return False
        if not isinstance(templates, dict):
            return False
        _DECKS.clear()
        _CARD_TEMPLATES.clear()
        for k, v in decks.items():
            if isinstance(v, dict) and "name" in v and "card_template_ids" in v:
                _DECKS[k] = {"name": str(v["name"]), "card_template_ids": list(v["card_template_ids"])}
        for k, v in templates.items():
            if isinstance(v, dict):
                _CARD_TEMPLATES[k] = {
                    "title": str(v.get("title", "")),
                    "body": str(v.get("body", "")),
                    "image_url": v.get("image_url") if v.get("image_url") is not None else None,
                    "back_text": v.get("back_text") if v.get("back_text") is not None else None,
                }
        # Persisted tile_metadata (deck_id per tile) for draw-card-for-tile
        tile_meta = data.get("tile_metadata")
        if isinstance(tile_meta, dict):
            for sym, meta in tile_meta.items():
                if len(sym) != 1 or not isinstance(meta, dict):
                    continue
                set_tile_metadata(
                    sym,
                    category=meta.get("category"),
                    difficulty=meta.get("difficulty"),
                    deck_id=meta.get("deck_id"),
                )
        return True
    except Exception:
        return False


def _save_decks_to_disk() -> None:
    """Atomically write current decks, card_templates, and tile_metadata to decks.json."""
    path = _get_decks_data_path() / "decks.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    data = {
        "decks": _DECKS,
        "card_templates": _CARD_TEMPLATES,
        "tile_metadata": dict(TILE_METADATA),
    }
    tmp = path.parent / (path.name + ".tmp")
    tmp.write_text(json.dumps(data, indent=2), encoding="utf-8")
    tmp.replace(path)


def _ensure_decks_loaded() -> None:
    """Load from disk if present, otherwise seed defaults. Idempotent after first call."""
    global _DECKS_INITIALIZED
    if _DECKS_INITIALIZED:
        return
    if _load_decks_from_disk():
        _DECKS_INITIALIZED = True
        return
    _seed_decks()
    _DECKS_INITIALIZED = True


def _seed_decks() -> None:
    if _DECKS:
        return
    for sym, name in TILE_NAMES.items():
        if sym in (".", "G", "1", "2", "3", "4", "C") or not name:
            continue
        deck_id = re.sub(r"[^a-z0-9_]", "_", name.lower()).strip("_") or f"deck_{sym}"
        if deck_id not in _DECKS:
            _DECKS[deck_id] = {"name": f"{name} Deck", "card_template_ids": []}
    if "default" not in _DECKS:
        _DECKS["default"] = {"name": "Default", "card_template_ids": []}


def _list_decks() -> list[DeckEntry]:
    _ensure_decks_loaded()
    return [DeckEntry(id=k, name=v["name"], card_template_ids=v["card_template_ids"]) for k, v in _DECKS.items()]


def get_deck_by_id(deck_id: str) -> DeckEntry | None:
    """Return deck by id for use by other routes (e.g. card export)."""
    _ensure_decks_loaded()
    if deck_id not in _DECKS:
        return None
    d = _DECKS[deck_id]
    return DeckEntry(id=deck_id, name=d["name"], card_template_ids=d["card_template_ids"])


def get_template_by_id(template_id: str) -> CardTemplateEntry | None:
    """Return card template by id for use by other routes (e.g. card export)."""
    if template_id not in _CARD_TEMPLATES:
        return None
    t = _CARD_TEMPLATES[template_id]
    return CardTemplateEntry(
        id=template_id,
        title=t.get("title", ""),
        body=t.get("body", ""),
        image_url=t.get("image_url"),
        back_text=t.get("back_text"),
    )


def get_templates_by_ids(template_ids: list[str]) -> list[CardTemplateEntry]:
    """Return card template entries for the given ids (missing ids yield placeholder entry)."""
    result: list[CardTemplateEntry] = []
    for tid in template_ids:
        t = get_template_by_id(tid)
        if t is not None:
            result.append(t)
        else:
            result.append(CardTemplateEntry(id=tid, title=tid, body="", image_url=None, back_text=None))
    return result


@router.get("", response_model=ConfigResponse)
def get_config() -> ConfigResponse:
    """Return presets, tile metadata, decks, and option schema for the UI."""
    _ensure_decks_loaded()
    tile_styles = {
        sym: TileStyle(fg=s["fg"], bg=s["bg"], glyph=s["glyph"])
        for sym, s in TILE_STYLES.items()
    }
    tile_metadata = {sym: dict(get_tile_metadata(sym)) for sym in TILE_COLORS}
    return ConfigResponse(
        board_presets=[list(p) for p in BOARD_PRESETS],
        tileset_presets={k: dict(v) for k, v in TILESET_PRESETS.items()},
        tile_names=dict(TILE_NAMES),
        tile_colors=dict(TILE_COLORS),
        tile_styles=tile_styles,
        tile_rules=dict(TILE_RULES),
        tile_metadata=tile_metadata,
        decks=_list_decks(),
        board_size_min=BOARD_SIZE_MIN,
        board_size_max=BOARD_SIZE_MAX,
        symmetry_choices=["none", "horizontal", "vertical", "both"],
        goal_placement_choices=["center", "random"],
        start_placement_choices=["corners", "random"],
        generation_mode_choices=list(GENERATION_MODE_CHOICES),
    )


@router.patch("/tile-metadata")
def patch_tile_metadata(req: TileMetadataPatchRequest) -> dict[str, str]:
    """Update tile metadata (category, difficulty, deck_id) for given symbols and persist to disk."""
    for symbol, update in req.tile_metadata.items():
        if len(symbol) != 1:
            continue
        set_tile_metadata(
            symbol,
            category=update.category,
            difficulty=update.difficulty,
            deck_id=update.deck_id,
        )
    _save_decks_to_disk()
    return {"status": "ok"}


# ---------------------------------------------------------------------------
# Decks CRUD
# ---------------------------------------------------------------------------


@router.get("/decks", response_model=list[DeckEntry])
def list_decks() -> list[DeckEntry]:
    """List all decks (for tile–deck mapping and card management)."""
    return _list_decks()


def _normalize_deck_id(deck_id: str) -> str:
    """Normalize to match seed_decks / board_core deck id format."""
    return re.sub(r"[^a-z0-9_]", "_", (deck_id or "").lower()).strip("_") or "default"


@router.get("/decks/{deck_id}/draw", response_model=CardTemplateEntry)
def draw_card(deck_id: str) -> CardTemplateEntry:
    """Draw a random card from the deck. Returns 404 if deck not found or deck has no cards."""
    _ensure_decks_loaded()
    deck = get_deck_by_id(deck_id)
    if deck is None:
        deck = get_deck_by_id(_normalize_deck_id(deck_id))
    if deck is None or not deck.card_template_ids:
        raise HTTPException(status_code=404, detail="Deck not found or has no cards")
    template_id = random.choice(deck.card_template_ids)
    card = get_template_by_id(template_id)
    if card is None:
        raise HTTPException(status_code=404, detail="Card template not found")
    return card


def _placeholder_card(title: str = "Nothing happens", body: str = "No card drawn for this tile.") -> CardTemplateEntry:
    """Return a placeholder card so the UI can always show the modal when landing on a special tile."""
    return CardTemplateEntry(
        id="__placeholder",
        title=title,
        body=body,
        image_url=None,
        back_text=None,
    )


@router.get("/draw-card-for-tile", response_model=CardTemplateEntry)
def draw_card_for_tile(symbol: str) -> CardTemplateEntry:
    """Draw a random card for the given tile symbol. Uses tile metadata deck_id. Returns placeholder if no deck or deck empty."""
    if not symbol or len(symbol) != 1:
        raise HTTPException(status_code=400, detail="symbol must be a single character")
    _ensure_decks_loaded()
    meta = get_tile_metadata(symbol)
    deck_id = (meta.get("deck_id") or "").strip()
    if not deck_id:
        return _placeholder_card(body="This tile has no deck assigned.")
    deck = get_deck_by_id(deck_id)
    if deck is None:
        deck = get_deck_by_id(_normalize_deck_id(deck_id))
    if deck is None or not deck.card_template_ids:
        return _placeholder_card(body="This deck has no cards.")
    template_id = random.choice(deck.card_template_ids)
    card = get_template_by_id(template_id)
    if card is None:
        return _placeholder_card(body="Card not found.")
    return card


@router.post("/decks", response_model=DeckEntry)
def create_deck(req: DeckCreateRequest) -> DeckEntry:
    """Create a new deck."""
    _ensure_decks_loaded()
    deck_id = re.sub(r"[^a-z0-9_]", "_", req.name.lower()).strip("_") or "deck"
    if not deck_id:
        deck_id = "deck"
    base = deck_id
    n = 0
    while deck_id in _DECKS:
        n += 1
        deck_id = f"{base}_{n}"
    _DECKS[deck_id] = {"name": req.name, "card_template_ids": list(req.card_template_ids)}
    _save_decks_to_disk()
    return DeckEntry(id=deck_id, name=_DECKS[deck_id]["name"], card_template_ids=_DECKS[deck_id]["card_template_ids"])


@router.patch("/decks/{deck_id}", response_model=DeckEntry)
def update_deck(deck_id: str, req: DeckUpdateRequest) -> DeckEntry:
    """Update a deck's name and/or card template list."""
    _ensure_decks_loaded()
    if deck_id not in _DECKS:
        raise HTTPException(status_code=404, detail="Deck not found")
    if req.name is not None:
        _DECKS[deck_id]["name"] = req.name
    if req.card_template_ids is not None:
        _DECKS[deck_id]["card_template_ids"] = list(req.card_template_ids)
    d = _DECKS[deck_id]
    _save_decks_to_disk()
    return DeckEntry(id=deck_id, name=d["name"], card_template_ids=d["card_template_ids"])


@router.delete("/decks/{deck_id}")
def delete_deck(deck_id: str) -> dict[str, str]:
    """Delete a deck (tile metadata referencing it may keep the id as-is)."""
    _ensure_decks_loaded()
    if deck_id not in _DECKS:
        raise HTTPException(status_code=404, detail="Deck not found")
    del _DECKS[deck_id]
    _save_decks_to_disk()
    return {"status": "ok"}


def _normalize_header(h: str) -> str:
    return (h or "").strip().lower()


def _deck_id_from_name(name: str) -> str:
    deck_id = re.sub(r"[^a-z0-9_]", "_", (name or "").lower()).strip("_") or "deck"
    return deck_id if deck_id else "deck"


def _ensure_deck_exists(deck_name: str) -> tuple[str, bool]:
    """Return (deck_id, True if newly created)."""
    deck_id = _deck_id_from_name(deck_name)
    base = deck_id
    n = 0
    while deck_id in _DECKS:
        existing_name = _DECKS[deck_id]["name"]
        if existing_name == deck_name:
            return deck_id, False
        n += 1
        deck_id = f"{base}_{n}"
    _DECKS[deck_id] = {"name": deck_name, "card_template_ids": []}
    return deck_id, True


def _parse_import_rows(
    rows: list[list[Any]],
) -> list[tuple[str, str, str, str | None, str | None]]:
    """Parse rows with header in first row. Return list of (deck_name, title, body, image_url, back_text)."""
    if not rows:
        raise HTTPException(status_code=400, detail="File has no rows")
    headers = [str(cell).strip() for cell in rows[0]]
    norm = [_normalize_header(h) for h in headers]
    deck_col = None
    title_col = None
    body_col = None
    image_col = None
    back_col = None
    for i, n in enumerate(norm):
        if n in ("deck", "deck name"):
            deck_col = i
        elif n == "title":
            title_col = i
        elif n == "body":
            body_col = i
        elif n in ("image url", "image_url"):
            image_col = i
        elif n in ("back text", "back_text"):
            back_col = i
    if deck_col is None:
        raise HTTPException(
            status_code=400,
            detail="Missing required column: Deck or Deck name",
        )
    if title_col is None:
        raise HTTPException(status_code=400, detail="Missing required column: Title")
    result: list[tuple[str, str, str, str | None, str | None]] = []
    for row in rows[1:]:
        deck_name = (row[deck_col] if deck_col < len(row) else "").strip()
        title = (row[title_col] if title_col < len(row) else "").strip()
        if not deck_name or not title:
            continue
        body = (row[body_col] if body_col is not None and body_col < len(row) else "").strip() or ""
        image_url = None
        if image_col is not None and image_col < len(row):
            v = (row[image_col] or "").strip()
            image_url = v if v else None
        back_text = None
        if back_col is not None and back_col < len(row):
            v = (row[back_col] or "").strip()
            back_text = v if v else None
        result.append((deck_name, title, body, image_url, back_text))
    return result


@router.post("/decks/import", response_model=DeckImportResult)
async def import_decks_from_file(file: UploadFile = File(..., alias="file")) -> DeckImportResult:
    """Import decks and cards from CSV or Excel (.xlsx). Columns: Deck, Title, Body, Image URL, Back text."""
    _ensure_decks_loaded()
    filename = (file.filename or "").lower()
    if not filename.endswith(".csv") and not filename.endswith(".xlsx"):
        raise HTTPException(
            status_code=400,
            detail="File must be CSV (.csv) or Excel (.xlsx)",
        )
    content = await file.read()
    rows: list[list[Any]] = []
    if filename.endswith(".csv"):
        try:
            text = content.decode("utf-8-sig")
            reader = csv.reader(io.StringIO(text))
            rows = list(reader)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid CSV: {e!s}") from e
    else:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            sheet = wb.active
            if sheet is None:
                raise HTTPException(status_code=400, detail="Excel file has no sheet")
            for row in sheet.iter_rows(values_only=True):
                rows.append(list(row) if row else [])
            wb.close()
        except ImportError:
            raise HTTPException(status_code=500, detail="Excel support requires openpyxl") from None
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e!s}") from e
    if not rows:
        raise HTTPException(status_code=400, detail="File has no data rows")
    parsed = _parse_import_rows(rows)
    new_deck_ids: set[str] = set()
    updated_deck_ids: set[str] = set()
    cards_created = 0
    for deck_name, title, body, image_url, back_text in parsed:
        deck_id, created = _ensure_deck_exists(deck_name)
        if created:
            new_deck_ids.add(deck_id)
        else:
            updated_deck_ids.add(deck_id)
        template_id = _next_template_id()
        _CARD_TEMPLATES[template_id] = {
            "title": title,
            "body": body,
            "image_url": image_url,
            "back_text": back_text,
        }
        _DECKS[deck_id]["card_template_ids"].append(template_id)
        cards_created += 1
    _save_decks_to_disk()
    return DeckImportResult(
        decks_created=len(new_deck_ids),
        decks_updated=len(updated_deck_ids),
        cards_created=cards_created,
    )


# ---------------------------------------------------------------------------
# Card templates CRUD
# ---------------------------------------------------------------------------


def _next_template_id() -> str:
    n = len(_CARD_TEMPLATES)
    base = "card"
    candidate = f"{base}_{n}"
    while candidate in _CARD_TEMPLATES:
        n += 1
        candidate = f"{base}_{n}"
    return candidate


@router.get("/card-templates", response_model=list[CardTemplateEntry])
def list_card_templates(deck_id: str | None = None) -> list[CardTemplateEntry]:
    """List all card templates, or only those in the given deck (by resolving card_template_ids)."""
    _ensure_decks_loaded()
    if deck_id is not None:
        deck = get_deck_by_id(deck_id)
        if deck is None:
            return []
        return get_templates_by_ids(deck.card_template_ids)
    return [
        CardTemplateEntry(
            id=k,
            title=v.get("title", ""),
            body=v.get("body", ""),
            image_url=v.get("image_url"),
            back_text=v.get("back_text"),
        )
        for k, v in _CARD_TEMPLATES.items()
    ]


@router.post("/card-templates", response_model=CardTemplateEntry)
def create_card_template(req: CardTemplateCreateRequest) -> CardTemplateEntry:
    """Create a new card template."""
    _ensure_decks_loaded()
    template_id = _next_template_id()
    _CARD_TEMPLATES[template_id] = {
        "title": req.title,
        "body": req.body,
        "image_url": req.image_url,
        "back_text": req.back_text,
    }
    t = _CARD_TEMPLATES[template_id]
    _save_decks_to_disk()
    return CardTemplateEntry(
        id=template_id,
        title=t["title"],
        body=t["body"],
        image_url=t.get("image_url"),
        back_text=t.get("back_text"),
    )


@router.patch("/card-templates/{template_id}", response_model=CardTemplateEntry)
def update_card_template(template_id: str, req: CardTemplateUpdateRequest) -> CardTemplateEntry:
    """Update a card template."""
    _ensure_decks_loaded()
    if template_id not in _CARD_TEMPLATES:
        raise HTTPException(status_code=404, detail="Card template not found")
    t = _CARD_TEMPLATES[template_id]
    if req.title is not None:
        t["title"] = req.title
    if req.body is not None:
        t["body"] = req.body
    if req.image_url is not None:
        t["image_url"] = req.image_url
    if req.back_text is not None:
        t["back_text"] = req.back_text
    _save_decks_to_disk()
    return CardTemplateEntry(
        id=template_id,
        title=t.get("title", ""),
        body=t.get("body", ""),
        image_url=t.get("image_url"),
        back_text=t.get("back_text"),
    )


@router.delete("/card-templates/{template_id}")
def delete_card_template(template_id: str) -> dict[str, str]:
    """Delete a card template and remove its id from any deck's card_template_ids."""
    _ensure_decks_loaded()
    if template_id not in _CARD_TEMPLATES:
        raise HTTPException(status_code=404, detail="Card template not found")
    del _CARD_TEMPLATES[template_id]
    for deck_id, d in list(_DECKS.items()):
        ids = d.get("card_template_ids", [])
        if template_id in ids:
            d["card_template_ids"] = [i for i in ids if i != template_id]
    _save_decks_to_disk()
    return {"status": "ok"}
